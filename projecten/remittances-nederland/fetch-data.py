#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.14"
# dependencies = [
#     "click==8.5.0",
#     "niquests==3.21.1",
#     "openpyxl==3.1.5",
#     "orjson==3.12.0",
#     "pydantic==2.13.5",
#     "pytest==9.1.1",
#     "pytest-cov==7.1.0",
#     "structlog==26.1.0",
# ]
# ///

"""Build the source-backed dataset for the Dutch Remittance Monitor."""

import contextlib
import datetime as dt
import io
import os
import subprocess as sp
import sys
import tempfile
from collections.abc import Iterable
from pathlib import Path
from typing import Any, TypedDict

import click
import niquests as http
import openpyxl as op
import orjson as json
import pytest
import structlog as sl
import structlog.stdlib as log
from click.testing import CliRunner
from pydantic import BaseModel, ConfigDict, Field

logger = log.get_logger(__name__)

CBS_BASE_URL = "https://datasets.cbs.nl/odata/v1/CBS/86136NED"
CBS_NATIONAL_ACCOUNTS_URL = "https://opendata.cbs.nl/ODataApi/OData/85879NED"
DNB_WORKBOOK_URL = (
    "https://www.dnb.nl/media/vkncb1xg/"
    "betalingsbalans-data-op-maat-vraag-remittances.xlsx"
)
VAT_RATES_URL = (
    "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/"
    "belastingdienst/zakelijk/btw/btw_berekenen_aan_uw_klanten/"
    "btw_berekenen/btw_tarief/"
)
CBS_PERIOD = "2024JJ00"
TOTAL_SEX = "T001038"
TOTAL_ORIGIN_COUNTRY = "T001040"
HTTP_HEADERS = {
    "User-Agent": "tbb-remittance-monitor/1.0 (+https://tbataafschebroederschap.nl)",
    "Accept": "*/*",
}

MEASURES = {
    "population": "M006774_1",
    "standardized_income": "M000222",
    "income_recipients": "M006774_2",
    "personal_income": "M001974",
    "economically_independent": "M006771",
    "poverty": "M006777",
}

REFERENCE_GROUPS = {
    "all-residents": ("T001638", "Alle inwoners"),
    "born-nl-parents-nl": ("A051760", "Geboren in Nederland, ouders in Nederland"),
    "born-nl-one-parent-nl": (
        "A051777",
        "Geboren in Nederland, één ouder in Nederland",
    ),
    "born-nl-both-parents-abroad": (
        "A051778",
        "Geboren in Nederland, beide ouders buiten Nederland",
    ),
    "born-abroad": ("A051736", "Geboren buiten Nederland"),
}

COUNTRY_GROUPS = {
    "belgie": ("H008552", "België"),
    "duitsland": ("H008592", "Duitsland"),
    "verenigd-koninkrijk": ("H008776", "Verenigd Koninkrijk"),
    "bulgarije": ("H008567", "Bulgarije"),
    "polen": ("H008718", "Polen"),
    "roemenie": ("H008723", "Roemenië"),
    "indonesie": ("H008632", "Indonesië"),
    "marokko": ("H008673", "Marokko"),
    "nederlands-caribisch-gebied": ("H007119", "Nederlands-Caribisch gebied"),
    "suriname": ("H008751", "Suriname"),
    "turkije": ("H008766", "Turkije"),
    "afghanistan": ("H008533", "Afghanistan"),
    "eritrea": ("H008597", "Eritrea"),
    "irak": ("H008633", "Irak"),
    "iran": ("H008634", "Iran"),
    "somalie": ("H008747", "Somalië"),
    "syrie": ("H008753", "Syrië"),
}

CORRIDOR_MATCHES = {
    "Turkije": "turkije",
    "Marokko": "marokko",
    "Suriname": "suriname",
    "Indonesië": "indonesie",
    "Roemenië": "roemenie",
    "Bulgarije": "bulgarije",
    "Polen": "polen",
    "Duitsland": "duitsland",
    "België": "belgie",
}

DUTCH_DESTINATION_LABELS = {
    "Turkey": "Turkije",
    "Morocco": "Marokko",
    "Suriname": "Suriname",
    "Indonesia": "Indonesië",
    "Romania": "Roemenië",
    "Bulgaria": "Bulgarije",
    "Philippines": "Filipijnen",
    "Spain": "Spanje",
    "Brazil": "Brazilië",
    "Ghana": "Ghana",
    "Colombia": "Colombia",
    "India": "India",
    "Poland": "Polen",
    "Curaçao": "Curaçao",
    "Germany": "Duitsland",
    "Pakistan": "Pakistan",
    "Dominican Republic": "Dominicaanse Republiek",
    "Greece": "Griekenland",
    "Cape Verde": "Kaapverdië",
    "China": "China",
    "Egypt": "Egypte",
    "France": "Frankrijk",
    "Italy": "Italië",
    "Belgium": "België",
    "Serbia": "Servië",
}


class InputError(Exception):
    """Report an invalid or changed source dataset."""


class CountryComponent(TypedDict):
    """Represent one birthplace component of an origin-country group."""

    label: str
    populationThousands: float
    incomeRecipientsThousands: float
    averagePersonalIncomeThousandEur: float
    totalPersonalIncomeMillionEur: float
    averageStandardizedIncomeThousandEur: float
    economicallyIndependentPct: float | None
    povertyPct: float | None


class Source(BaseModel):
    """Describe an upstream source and its snapshot."""

    id: str
    title: str
    url: str
    publisher: str
    retrieved_at: str = Field(alias="retrievedAt")
    period: str

    model_config = ConfigDict(populate_by_name=True)


class IncomeGroup(BaseModel):
    """Store comparable CBS income values for one population group."""

    id: str
    label: str
    kind: str
    population_thousands: float = Field(alias="populationThousands")
    income_recipients_thousands: float = Field(alias="incomeRecipientsThousands")
    average_personal_income_thousand_eur: float = Field(
        alias="averagePersonalIncomeThousandEur"
    )
    personal_income_per_member_thousand_eur: float = Field(
        alias="personalIncomePerMemberThousandEur"
    )
    average_standardized_income_thousand_eur: float = Field(
        alias="averageStandardizedIncomeThousandEur"
    )
    total_personal_income_million_eur: float = Field(
        alias="totalPersonalIncomeMillionEur"
    )
    income_recipient_share_pct: float = Field(alias="incomeRecipientSharePct")
    economically_independent_pct: float | None = Field(
        alias="economicallyIndependentPct"
    )
    poverty_pct: float | None = Field(alias="povertyPct")
    birth_breakdown: list[dict[str, Any]] = Field(
        default_factory=list, alias="birthBreakdown"
    )

    model_config = ConfigDict(populate_by_name=True)


class Corridor(BaseModel):
    """Store a DNB remittance destination and its annual amounts."""

    id: str
    destination: str
    destination_source: str = Field(alias="destinationSource")
    values: list[dict[str, int | float]]

    model_config = ConfigDict(populate_by_name=True)


class Dataset(BaseModel):
    """Validate the published view model before it is written."""

    schema_version: int = Field(alias="schemaVersion")
    generated_at: str = Field(alias="generatedAt")
    snapshot: dict[str, Any]
    definitions: dict[str, str]
    sources: list[Source]
    remittances: dict[str, Any]
    income: dict[str, Any]
    economic_context: dict[str, Any] = Field(alias="economicContext")
    ecological_comparisons: list[dict[str, Any]] = Field(alias="ecologicalComparisons")
    quality: dict[str, Any]

    model_config = ConfigDict(populate_by_name=True)


def configure_logging() -> None:
    """Send structured diagnostic logs to stderr."""
    sl.configure(
        processors=[
            sl.processors.TimeStamper(fmt="iso", utc=True),
            sl.processors.add_log_level,
            sl.dev.ConsoleRenderer(colors=sys.stderr.isatty()),
        ],
        wrapper_class=sl.make_filtering_bound_logger("debug"),
        logger_factory=sl.PrintLoggerFactory(file=sys.stderr),
        cache_logger_on_first_use=False,
    )


def fetch_json_rows(
    url: str, params: dict[str, str] | None = None
) -> list[dict[str, Any]]:
    """Fetch all rows from a paginated OData endpoint."""
    rows: list[dict[str, Any]] = []
    next_url: str | None = url
    next_params = params
    while next_url:
        response = http.get(next_url, params=next_params, timeout=60)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict) or not isinstance(payload.get("value"), list):
            raise InputError(f"unexpected OData response from {next_url}")
        rows.extend(payload["value"])
        next_url = payload.get("@odata.nextLink") or payload.get("odata.nextLink")
        next_params = None
    return rows


def fetch_binary(url: str) -> bytes:
    """Download a binary source with an explicit timeout."""
    response = http.get(url, headers=HTTP_HEADERS, timeout=60)
    response.raise_for_status()
    content = response.content
    if content is None:
        raise InputError(f"empty binary response from {url}")
    return content


def fetch_cbs_observations() -> list[dict[str, Any]]:
    """Fetch the 2024 all-sex CBS observations used by the monitor."""
    return fetch_json_rows(
        f"{CBS_BASE_URL}/Observations",
        {
            "$filter": (f"Geslacht eq '{TOTAL_SEX}' and Perioden eq '{CBS_PERIOD}'"),
            "$top": "10000",
        },
    )


def fetch_national_accounts() -> list[dict[str, Any]]:
    """Fetch aligned 2024 GDP and household-consumption denominators."""
    return fetch_json_rows(
        f"{CBS_NATIONAL_ACCOUNTS_URL}/TypedDataSet",
        {
            "$filter": ("SoortGegevens eq 'A045297' and Perioden eq '2024JJ00'"),
            "$select": (
                "SoortGegevens,Perioden,BrutoBinnenlandsProduct_2,Huishoudens_9"
            ),
        },
    )


def numeric(value: Any, *, context: str) -> float:
    """Return a numeric source value or fail with a useful location."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise InputError(f"missing numeric value for {context}")
    return float(value)


def observation_index(
    rows: Iterable[dict[str, Any]],
) -> dict[tuple[str, str, str], float]:
    """Index CBS long-form observations by origin, origin country and measure."""
    index: dict[tuple[str, str, str], float] = {}
    for row in rows:
        measure_value = row.get("Measure")
        if not isinstance(measure_value, str) or measure_value not in MEASURES.values():
            continue
        if row.get("Value") is None:
            continue
        measure = measure_value
        key = (str(row.get("Herkomst")), str(row.get("Herkomstland")), measure)
        index[key] = numeric(row.get("Value"), context="/".join(key))
    return index


def lookup(
    index: dict[tuple[str, str, str], float],
    origin_code: str,
    country_code: str,
    measure: str,
) -> float:
    """Read one required CBS value from the observation index."""
    key = (origin_code, country_code, MEASURES[measure])
    if key not in index:
        raise InputError(f"CBS observation missing: {'/'.join(key)}")
    return index[key]


def lookup_optional(
    index: dict[tuple[str, str, str], float],
    origin_code: str,
    country_code: str,
    measure: str,
) -> float | None:
    """Read a CBS value that may be suppressed for a small subgroup."""
    return index.get((origin_code, country_code, MEASURES[measure]))


def make_income_group(
    *,
    group_id: str,
    label: str,
    kind: str,
    population: float,
    recipients: float,
    average_personal: float,
    standardized: float,
    independent: float | None,
    poverty: float | None,
    birth_breakdown: list[dict[str, Any]] | None = None,
) -> IncomeGroup:
    """Derive totals and per-member metrics with explicit units."""
    if population <= 0 or recipients <= 0:
        raise InputError(f"non-positive CBS denominator for {group_id}")
    total_income = recipients * average_personal
    return IncomeGroup(
        id=group_id,
        label=label,
        kind=kind,
        populationThousands=round(population, 1),
        incomeRecipientsThousands=round(recipients, 1),
        averagePersonalIncomeThousandEur=round(average_personal, 1),
        personalIncomePerMemberThousandEur=round(total_income / population, 1),
        averageStandardizedIncomeThousandEur=round(standardized, 1),
        totalPersonalIncomeMillionEur=round(total_income, 1),
        incomeRecipientSharePct=round(recipients / population * 100, 1),
        economicallyIndependentPct=(
            None if independent is None else round(independent, 1)
        ),
        povertyPct=None if poverty is None else round(poverty, 1),
        birthBreakdown=birth_breakdown or [],
    )


def build_reference_groups(
    index: dict[tuple[str, str, str], float],
) -> list[IncomeGroup]:
    """Build benchmark and origin-generation CBS groups."""
    result = []
    for group_id, (origin_code, label) in REFERENCE_GROUPS.items():
        result.append(
            make_income_group(
                group_id=group_id,
                label=label,
                kind=(
                    "reference" if group_id == "born-nl-parents-nl" else "generation"
                ),
                population=lookup(
                    index, origin_code, TOTAL_ORIGIN_COUNTRY, "population"
                ),
                recipients=lookup(
                    index, origin_code, TOTAL_ORIGIN_COUNTRY, "income_recipients"
                ),
                average_personal=lookup(
                    index, origin_code, TOTAL_ORIGIN_COUNTRY, "personal_income"
                ),
                standardized=lookup(
                    index, origin_code, TOTAL_ORIGIN_COUNTRY, "standardized_income"
                ),
                independent=lookup(
                    index,
                    origin_code,
                    TOTAL_ORIGIN_COUNTRY,
                    "economically_independent",
                ),
                poverty=lookup(index, origin_code, TOTAL_ORIGIN_COUNTRY, "poverty"),
            )
        )
    return result


def build_country_groups(
    index: dict[tuple[str, str, str], float],
) -> list[IncomeGroup]:
    """Aggregate each origin country over Netherlands- and foreign-born."""
    result = []
    for group_id, (country_code, label) in COUNTRY_GROUPS.items():
        components: list[CountryComponent] = []
        for origin_code, birth_label in (
            ("A051735", "Geboren in Nederland"),
            ("A051736", "Geboren buiten Nederland"),
        ):
            population = lookup(index, origin_code, country_code, "population")
            recipients = lookup(index, origin_code, country_code, "income_recipients")
            average_personal = lookup(
                index, origin_code, country_code, "personal_income"
            )
            components.append(
                {
                    "label": birth_label,
                    "populationThousands": population,
                    "incomeRecipientsThousands": recipients,
                    "averagePersonalIncomeThousandEur": average_personal,
                    "totalPersonalIncomeMillionEur": recipients * average_personal,
                    "averageStandardizedIncomeThousandEur": lookup(
                        index, origin_code, country_code, "standardized_income"
                    ),
                    "economicallyIndependentPct": lookup_optional(
                        index,
                        origin_code,
                        country_code,
                        "economically_independent",
                    ),
                    "povertyPct": lookup_optional(
                        index, origin_code, country_code, "poverty"
                    ),
                }
            )

        population_total = sum(item["populationThousands"] for item in components)
        recipients_total = sum(item["incomeRecipientsThousands"] for item in components)
        income_total = sum(item["totalPersonalIncomeMillionEur"] for item in components)

        standardized = weighted_component_value(
            components,
            [item["averageStandardizedIncomeThousandEur"] for item in components],
            population_total,
        )
        if standardized is None:
            raise InputError(f"standardized income missing for {group_id}")

        result.append(
            make_income_group(
                group_id=group_id,
                label=label,
                kind="origin-country",
                population=population_total,
                recipients=recipients_total,
                average_personal=income_total / recipients_total,
                standardized=standardized,
                independent=weighted_component_value(
                    components,
                    [item["economicallyIndependentPct"] for item in components],
                    population_total,
                ),
                poverty=weighted_component_value(
                    components,
                    [item["povertyPct"] for item in components],
                    population_total,
                ),
                birth_breakdown=[
                    {
                        key: round(value, 1) if isinstance(value, float) else value
                        for key, value in item.items()
                    }
                    for item in components
                ],
            )
        )
    return result


def weighted_component_value(
    components: list[CountryComponent],
    values: list[float | None],
    population_total: float,
) -> float | None:
    """Calculate a population-weighted component metric."""
    if any(value is None for value in values):
        return None
    numeric_values = [value for value in values if value is not None]
    return (
        sum(
            item["populationThousands"] * value
            for item, value in zip(components, numeric_values, strict=True)
        )
        / population_total
    )


def row_label(row: tuple[Any, ...], first_year_column: int) -> str:
    """Find the descriptive cell preceding the DNB year columns."""
    for cell in reversed(row[:first_year_column]):
        if isinstance(cell, str) and cell.strip():
            return cell.strip()
    return ""


def parse_dnb_workbook(
    content: bytes,
) -> tuple[list[int], list[Corridor], dict[int, float]]:
    """Parse the top-25 destinations and total remittances from DNB's workbook."""
    workbook = op.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = -1
    year_columns: dict[int, int] = {}
    for index, row in enumerate(rows):
        candidates = {
            column: int(value)
            for column, value in enumerate(row)
            if isinstance(value, (int, float)) and 2010 <= int(value) <= 2030
        }
        if len(candidates) >= 5:
            header_index = index
            year_columns = candidates
            break
    if header_index < 0:
        raise InputError("DNB workbook contains no recognizable year header")

    years = [year_columns[column] for column in sorted(year_columns)]
    first_year_column = min(year_columns)
    corridors: list[Corridor] = []
    totals: dict[int, float] = {}
    for row in rows[header_index + 1 :]:
        label = row_label(row, first_year_column)
        if not label or label in {"…", "..."}:
            continue
        values = {
            year: numeric(row[column], context=f"DNB/{label}/{year}")
            for column, year in year_columns.items()
            if column < len(row) and row[column] is not None
        }
        if len(values) != len(years):
            continue
        if label.casefold().startswith("totaal"):
            totals = values
            break
        corridor_id = (
            label.casefold().replace(" ", "-").replace("ç", "c").replace("ã", "a")
        )
        corridors.append(
            Corridor(
                id=corridor_id,
                destination=DUTCH_DESTINATION_LABELS.get(label, label),
                destinationSource=label,
                values=[{"year": year, "millionEur": values[year]} for year in years],
            )
        )
    if not totals:
        raise InputError("DNB workbook contains no total row")
    return years, corridors, totals


def value_for_year(corridor: Corridor, year: int) -> float:
    """Read one year's remittance value from a parsed corridor."""
    for item in corridor.values:
        if item["year"] == year:
            return float(item["millionEur"])
    raise InputError(f"corridor {corridor.destination_source} lacks {year}")


def validate_source_shapes(
    years: list[int], corridors: list[Corridor], totals: dict[int, float]
) -> None:
    """Fail loudly if a source update changes the published DNB contract."""
    if years != list(range(2016, 2026)):
        raise InputError(f"unexpected DNB year range: {years}")
    if len(corridors) != 25:
        raise InputError(f"expected 25 DNB corridors, received {len(corridors)}")
    if totals.get(2024) != 451 or totals.get(2025) != 478:
        raise InputError("DNB control totals changed for 2024 or 2025")
    missing = set(CORRIDOR_MATCHES) - {
        corridor.destination_source for corridor in corridors
    }
    if missing:
        raise InputError(f"DNB matched corridors missing: {sorted(missing)}")


def build_ecological_comparisons(
    country_groups: list[IncomeGroup],
    reference_groups: list[IncomeGroup],
    corridors: list[Corridor],
    netherlands_total_million_eur: float,
) -> list[dict[str, Any]]:
    """Join destination totals to same-named origin groups as an ecological proxy."""
    countries = {group.id: group for group in country_groups}
    destinations = {corridor.destination_source: corridor for corridor in corridors}
    reference = next(
        group for group in reference_groups if group.id == "born-nl-parents-nl"
    )
    result = []
    for destination, group_id in CORRIDOR_MATCHES.items():
        group = countries[group_id]
        remittance = value_for_year(destinations[destination], 2024)
        result.append(
            {
                "groupId": group.id,
                "group": group.label,
                "destination": destinations[destination].destination,
                "year": 2024,
                "remittanceMillionEur": remittance,
                "totalPersonalIncomeMillionEur": (
                    group.total_personal_income_million_eur
                ),
                "remittanceAsPctOfIncome": round(
                    remittance / group.total_personal_income_million_eur * 100,
                    2,
                ),
                "remittanceShareOfNetherlandsTotalPct": round(
                    remittance / netherlands_total_million_eur * 100,
                    1,
                ),
                "remittanceEurPerGroupMember": round(
                    remittance * 1000 / group.population_thousands
                ),
                "personalIncomePerMemberThousandEur": (
                    group.personal_income_per_member_thousand_eur
                ),
                "incomeVsReferencePct": round(
                    group.personal_income_per_member_thousand_eur
                    / reference.personal_income_per_member_thousand_eur
                    * 100,
                    1,
                ),
                "interpretation": ("corridorproxy; geen individuele afzendermeting"),
            }
        )
    return sorted(
        result,
        key=lambda item: item["remittanceAsPctOfIncome"],
        reverse=True,
    )


def build_economic_context(
    *,
    remittance_million_eur: float,
    reference_groups: list[IncomeGroup],
    national_accounts_rows: list[dict[str, Any]],
    comparisons: list[dict[str, Any]],
) -> dict[str, Any]:
    """Build aligned scale ratios and counterfactual scenario bounds."""
    if len(national_accounts_rows) != 1:
        raise InputError("expected exactly one 2024 CBS national-accounts observation")
    accounts = national_accounts_rows[0]
    gdp = numeric(
        accounts.get("BrutoBinnenlandsProduct_2"),
        context="CBS 85879NED/2024/GDP",
    )
    household_consumption = numeric(
        accounts.get("Huishoudens_9"),
        context="CBS 85879NED/2024/household consumption",
    )
    all_residents = next(
        group for group in reference_groups if group.id == "all-residents"
    )
    total_personal_income = all_residents.total_personal_income_million_eur
    matched_remittance = sum(
        numeric(
            comparison["remittanceMillionEur"],
            context=f"comparison/{comparison['groupId']}/remittance",
        )
        for comparison in comparisons
    )
    return {
        "year": 2024,
        "assessment": "causal economic damage is not identifiable",
        "grossRemittanceMillionEur": remittance_million_eur,
        "gdpMillionEur": gdp,
        "householdConsumptionMillionEur": household_consumption,
        "totalPersonalIncomeMillionEur": total_personal_income,
        "remittanceAsPctOfGdp": round(remittance_million_eur / gdp * 100, 4),
        "remittanceAsPctOfHouseholdConsumption": round(
            remittance_million_eur / household_consumption * 100,
            4,
        ),
        "remittanceAsPctOfPersonalIncome": round(
            remittance_million_eur / total_personal_income * 100,
            4,
        ),
        "matchedCorridorsMillionEur": round(matched_remittance, 1),
        "matchedCorridorsShareOfTotalPct": round(
            matched_remittance / remittance_million_eur * 100,
            1,
        ),
        "scenarioBounds": {
            "displacedDomesticSpendingMinimumMillionEur": 0,
            "displacedDomesticSpendingMaximumMillionEur": remittance_million_eur,
            "vatIncludedAt9PctMaximumMillionEur": round(
                remittance_million_eur * 9 / 109,
                1,
            ),
            "vatIncludedAt21PctMaximumMillionEur": round(
                remittance_million_eur * 21 / 121,
                1,
            ),
        },
        "scenarioNote": (
            "The range is not an estimate of actual damage. It varies the "
            "unobserved share that would otherwise have been spent in the "
            "Netherlands; saving, imports and other uses can make the domestic "
            "effect smaller."
        ),
    }


def build_dataset(
    *,
    retrieved_at: dt.datetime | None = None,
    cbs_rows: list[dict[str, Any]] | None = None,
    dnb_content: bytes | None = None,
    national_accounts_rows: list[dict[str, Any]] | None = None,
) -> Dataset:
    """Fetch, reconcile and validate all source data for the dashboard."""
    timestamp = retrieved_at or dt.datetime.now(dt.UTC)
    rows = cbs_rows if cbs_rows is not None else fetch_cbs_observations()
    workbook = (
        dnb_content if dnb_content is not None else fetch_binary(DNB_WORKBOOK_URL)
    )
    accounts_rows = (
        national_accounts_rows
        if national_accounts_rows is not None
        else fetch_national_accounts()
    )
    index = observation_index(rows)
    reference_groups = build_reference_groups(index)
    country_groups = build_country_groups(index)
    years, corridors, totals = parse_dnb_workbook(workbook)
    validate_source_shapes(years, corridors, totals)
    comparisons = build_ecological_comparisons(
        country_groups,
        reference_groups,
        corridors,
        totals[2024],
    )
    economic_context = build_economic_context(
        remittance_million_eur=totals[2024],
        reference_groups=reference_groups,
        national_accounts_rows=accounts_rows,
        comparisons=comparisons,
    )
    top25_by_year = {
        str(year): round(
            sum(value_for_year(corridor, year) for corridor in corridors), 1
        )
        for year in years
    }
    iso_timestamp = timestamp.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    cumulative = round(sum(totals.values()), 1)
    return Dataset(
        schemaVersion=2,
        generatedAt=iso_timestamp,
        snapshot={
            "cbsYear": 2024,
            "cbsStatus": "voorlopig",
            "dnbLatestYear": max(years),
            "dnbPeriod": f"{min(years)}–{max(years)}",
        },
        definitions={
            "reference": (
                "Persoon geboren in Nederland van wie beide ouders in Nederland "
                "zijn geboren. Dit is een herkomstproxy, geen nationaliteit of "
                "etniciteit."
            ),
            "remittance": (
                "Uitgaande particuliere geldovermaking naar een bestemming buiten "
                "Nederland, zoals door DNB geraamd."
            ),
            "ecologicalProxy": (
                "Remittance naar een land gedeeld door het totale persoonlijke "
                "inkomen van de gelijknamige CBS-herkomstgroep. Bestemmingsland "
                "en afzendergroep zijn niet gelijk."
            ),
        },
        sources=[
            Source(
                id="cbs-86136ned",
                title=("CBS 86136NED — Inkomen van personen; persoonskenmerken"),
                url=CBS_BASE_URL,
                publisher="Centraal Bureau voor de Statistiek",
                retrievedAt=iso_timestamp,
                period="2024 (voorlopig)",
            ),
            Source(
                id="dnb-remittances-2026",
                title="DNB — Uitgaande overmakingen vanuit Nederland",
                url=DNB_WORKBOOK_URL,
                publisher="De Nederlandsche Bank",
                retrievedAt=iso_timestamp,
                period="2016–2025",
            ),
            Source(
                id="cbs-85879ned",
                title="CBS 85879NED — Bbp, productie en bestedingen",
                url=CBS_NATIONAL_ACCOUNTS_URL,
                publisher="Centraal Bureau voor de Statistiek",
                retrievedAt=iso_timestamp,
                period="2024 (werkelijke prijzen)",
            ),
            Source(
                id="belastingdienst-btw-tarieven-2026",
                title="Belastingdienst — Btw-tarieven 9% en 21%",
                url=VAT_RATES_URL,
                publisher="Belastingdienst",
                retrievedAt=iso_timestamp,
                period="geraadpleegd 2026",
            ),
        ],
        remittances={
            "unit": "miljoen euro",
            "timeline": [{"year": year, "millionEur": totals[year]} for year in years],
            "corridors": [corridor.model_dump(by_alias=True) for corridor in corridors],
            "summary": {
                "latestYear": max(years),
                "latestTotalMillionEur": totals[max(years)],
                "cumulativeMillionEur": cumulative,
                "top25ShareLatestPct": round(
                    top25_by_year[str(max(years))] / totals[max(years)] * 100,
                    1,
                ),
                "change2020Pct": round((totals[2020] / totals[2019] - 1) * 100, 1),
            },
            "top25TotalMillionEurByYear": top25_by_year,
        },
        income={
            "unit": "duizend euro per jaar",
            "referenceGroupId": "born-nl-parents-nl",
            "referenceGroups": [
                group.model_dump(by_alias=True) for group in reference_groups
            ],
            "countryGroups": [
                group.model_dump(by_alias=True) for group in country_groups
            ],
        },
        economicContext=economic_context,
        ecologicalComparisons=comparisons,
        quality={
            "status": "geschikt met beperkingen",
            "warnings": [
                (
                    "DNB meet bestemmingslanden, niet de herkomst of "
                    "nationaliteit van afzenders."
                ),
                (
                    "De remittance van de referentiegroep is onbekend en wordt "
                    "niet als nul behandeld."
                ),
                (
                    "CBS 2024 is voorlopig; inkomens zijn gemiddelden en geen "
                    "fiscale nettobijdragen."
                ),
                (
                    "Bruto uitstroom is geen gemeten economische schade; daarvoor "
                    "ontbreekt het tegenfeitelijke gebruik van het geld."
                ),
                (
                    "De daling in 2020 is een controlepunt: de DNB-tabel labelt "
                    "geen methodebreuk."
                ),
                ("DNB-bedragen per bestemming zijn afgerond op hele miljoenen euro."),
            ],
            "receipts": {
                "cbsObservationCount": len(rows),
                "incomeReferenceGroupCount": len(reference_groups),
                "incomeCountryGroupCount": len(country_groups),
                "dnbYearCount": len(years),
                "dnbCorridorCount": len(corridors),
                "exactEcologicalMatchCount": len(comparisons),
                "dnb2024ControlTotalMillionEur": totals[2024],
                "dnb2025ControlTotalMillionEur": totals[2025],
                "nationalAccountsObservationCount": len(accounts_rows),
            },
        },
    )


def serialize_dataset(dataset: Dataset) -> bytes:
    """Serialize deterministically for reviewable repository diffs."""
    return json.dumps(
        dataset.model_dump(by_alias=True, mode="json"),
        option=(json.OPT_INDENT_2 | json.OPT_SORT_KEYS | json.OPT_APPEND_NEWLINE),
    )


def write_atomic(path: Path, content: bytes) -> None:
    """Atomically replace one generated artifact."""
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}.", dir=path.parent
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def check_sources() -> None:
    """Verify that all upstream data endpoints remain reachable."""
    cbs_response = http.get(f"{CBS_BASE_URL}/MeasureCodes?$top=1", timeout=20)
    cbs_response.raise_for_status()
    accounts_rows = fetch_national_accounts()
    if len(accounts_rows) != 1:
        raise InputError("CBS national-accounts readiness row is missing")
    dnb_content = fetch_binary(DNB_WORKBOOK_URL)
    if not dnb_content:
        raise InputError("DNB workbook response is empty")


@click.group()
def cli() -> None:
    """Refresh and validate the Dutch Remittance Monitor dataset."""
    configure_logging()


@cli.command(name="refresh")
@click.option(
    "--output",
    type=click.Path(path_type=Path, dir_okay=False),
    default=Path(__file__).with_name("data.json"),
    show_default=True,
    help="Primary generated JSON file.",
)
@click.option(
    "--mirror-output",
    type=click.Path(path_type=Path, dir_okay=False),
    default=None,
    help="Optional second copy for the data-product pipeline.",
)
@click.option(
    "--dry-run",
    is_flag=True,
    help="Validate sources without writing files.",
)
@click.option(
    "--yes",
    is_flag=True,
    help="Write without an interactive confirmation.",
)
def refresh_command(
    output: Path,
    mirror_output: Path | None,
    dry_run: bool,
    yes: bool,
) -> None:
    """Fetch CBS and DNB, validate the join and write the view model."""
    try:
        dataset = build_dataset()
    except (InputError, OSError, ValueError) as error:
        raise click.ClickException(str(error)) from error
    payload = serialize_dataset(dataset)
    outputs = [output] + ([mirror_output] if mirror_output else [])
    summary = dataset.remittances["summary"]
    if dry_run:
        click.echo(
            json.dumps(
                {
                    "status": "valid",
                    "bytes": len(payload),
                    "outputs": [str(path) for path in outputs],
                    "latestMillionEur": (summary["latestTotalMillionEur"]),
                }
            ).decode()
        )
        return
    if not yes and not click.confirm(
        f"{len(outputs)} gegenereerde databestand(en) vervangen?"
    ):
        raise click.Abort()
    for path in outputs:
        write_atomic(path, payload)
    logger.info(
        "dataset_refreshed",
        outputs=[str(path) for path in outputs],
        bytes=len(payload),
    )
    click.echo("\n".join(str(path) for path in outputs))


@cli.command(name="check")
def check_command() -> None:
    """Verify runtime imports and upstream source reachability."""
    try:
        check_sources()
    except (InputError, OSError, ValueError) as error:
        raise click.ClickException(str(error)) from error
    click.echo("ok")


def compact_pytest_output(output: str) -> str:
    """Remove pytest-cov banners while preserving its useful report."""
    lines = []
    for line in output.splitlines():
        is_section_banner = (
            line.startswith("=") and line.endswith("=") and " tests coverage " in line
        )
        is_platform_banner = (
            line.startswith("_")
            and line.endswith("_")
            and " coverage: platform " in line
        )
        if not is_section_banner and not is_platform_banner:
            lines.append(line)
    return "\n".join(lines).strip() + "\n"


@click.command(name="unit-test")
def _embedded_unit_test_command() -> None:
    """Run embedded tests and report line and branch coverage."""
    with tempfile.TemporaryDirectory(prefix="python-cli-coverage-") as directory:
        coverage_config = Path(directory) / ".coveragerc"
        coverage_config.write_text(
            os.linesep.join(
                (
                    "[run]",
                    "patch = subprocess",
                    "include =",
                    f"    {Path(__file__).resolve().as_posix()}",
                    "",
                )
            ),
            encoding="utf-8",
        )
        previous_coverage_file = os.environ.get("COVERAGE_FILE")
        os.environ["COVERAGE_FILE"] = str(Path(directory) / ".coverage")
        pytest_output = io.StringIO()
        try:
            with contextlib.redirect_stdout(pytest_output):
                result = pytest.main(
                    [
                        "--cov",
                        "--cov-branch",
                        "--cov-config",
                        str(coverage_config),
                        "--cov-report=term-missing",
                        "-p",
                        "no:cacheprovider",
                        __file__,
                        "-q",
                    ]
                )
        finally:
            if previous_coverage_file is None:
                os.environ.pop("COVERAGE_FILE", None)
            else:
                os.environ["COVERAGE_FILE"] = previous_coverage_file
    click.echo(
        compact_pytest_output(pytest_output.getvalue()),
        nl=False,
    )
    raise SystemExit(result)


cli.add_command(_embedded_unit_test_command)


def fake_dnb_workbook() -> bytes:
    """Create a structurally representative workbook for parser tests."""
    workbook = op.Workbook()
    worksheet = workbook.active
    worksheet.append(["bron"])
    worksheet.append([None, "Land", *range(2016, 2026)])
    for index, destination in enumerate(
        DUTCH_DESTINATION_LABELS,
        start=1,
    ):
        worksheet.append([None, destination, *([index] * 10)])
    worksheet.append([None, "Totaal", *([478] * 10)])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_make_income_group_derives_totals() -> None:
    group = make_income_group(
        group_id="test",
        label="Test",
        kind="test",
        population=200,
        recipients=160,
        average_personal=40,
        standardized=42,
        independent=55,
        poverty=4,
    )
    assert group.total_personal_income_million_eur == 6400
    assert group.personal_income_per_member_thousand_eur == 32
    assert group.income_recipient_share_pct == 80


def test_build_economic_context_keeps_damage_as_scenario() -> None:
    all_residents = make_income_group(
        group_id="all-residents",
        label="Alle inwoners",
        kind="generation",
        population=17_507.3,
        recipients=14_480.6,
        average_personal=42.2,
        standardized=45.3,
        independent=70,
        poverty=3,
    )
    context = build_economic_context(
        remittance_million_eur=451,
        reference_groups=[all_residents],
        national_accounts_rows=[
            {
                "BrutoBinnenlandsProduct_2": 1_115_252,
                "Huishoudens_9": 487_810,
            }
        ],
        comparisons=[{"groupId": "test", "remittanceMillionEur": 248}],
    )
    assert context["assessment"] == "causal economic damage is not identifiable"
    assert context["remittanceAsPctOfGdp"] == 0.0404
    assert context["remittanceAsPctOfHouseholdConsumption"] == 0.0925
    assert context["scenarioBounds"]["vatIncludedAt21PctMaximumMillionEur"] == 78.3


def test_parse_dnb_workbook_reads_corridors_and_total() -> None:
    years, corridors, totals = parse_dnb_workbook(fake_dnb_workbook())
    assert years == list(range(2016, 2026))
    assert len(corridors) == 25
    assert corridors[0].destination == "Turkije"
    assert totals[2025] == 478


def test_lookup_rejects_missing_value() -> None:
    index = observation_index(
        [
            {
                "Herkomst": "a",
                "Herkomstland": "b",
                "Measure": MEASURES["population"],
                "Value": None,
            }
        ]
    )
    with pytest.raises(InputError, match="CBS observation missing"):
        lookup(index, "a", "b", "population")


def test_cli_help_shows_operational_commands() -> None:
    result = CliRunner().invoke(cli, ["--help"])
    assert result.exit_code == 0
    assert "refresh" in result.stdout
    assert "check" in result.stdout
    assert "unit-test" in result.stdout


def test_check_command_reports_readiness(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        sys.modules[__name__],
        "check_sources",
        lambda: None,
    )
    result = CliRunner().invoke(cli, ["check"])
    assert result.exit_code == 0
    assert result.stdout == "ok\n"


def test_refresh_dry_run_does_not_write(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    minimal = Dataset(
        schemaVersion=1,
        generatedAt="2026-01-01T00:00:00Z",
        snapshot={},
        definitions={},
        sources=[],
        remittances={"summary": {"latestTotalMillionEur": 478}},
        income={},
        economicContext={},
        ecologicalComparisons=[],
        quality={},
    )
    monkeypatch.setattr(
        sys.modules[__name__],
        "build_dataset",
        lambda: minimal,
    )
    with CliRunner().isolated_filesystem():
        result = CliRunner().invoke(
            cli,
            ["refresh", "--dry-run", "--output", "data.json"],
        )
        assert result.exit_code == 0
        assert not Path("data.json").exists()
        assert '"status":"valid"' in result.stdout


def test_refresh_writes_primary_and_mirror(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    minimal = Dataset(
        schemaVersion=1,
        generatedAt="2026-01-01T00:00:00Z",
        snapshot={},
        definitions={},
        sources=[],
        remittances={"summary": {"latestTotalMillionEur": 478}},
        income={},
        economicContext={},
        ecologicalComparisons=[],
        quality={},
    )
    monkeypatch.setattr(
        sys.modules[__name__],
        "build_dataset",
        lambda: minimal,
    )
    with CliRunner().isolated_filesystem():
        result = CliRunner().invoke(
            cli,
            [
                "refresh",
                "--yes",
                "--output",
                "data.json",
                "--mirror-output",
                "mirror/data.json",
            ],
        )
        assert result.exit_code == 0
        assert Path("data.json").read_bytes() == Path("mirror/data.json").read_bytes()


def test_subprocess_help_keeps_stdout_clean() -> None:
    result = sp.run(
        [sys.executable, __file__, "--help"],
        check=False,
        capture_output=True,
        text=True,
        timeout=10,
    )
    assert result.returncode == 0
    assert "Usage:" in result.stdout


if __name__ == "__main__":
    cli()
